from django.http import HttpResponse
from openpyxl import Workbook
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAdminUser
from rest_framework.generics import ListAPIView
from rest_framework.views import APIView

from .serializers import FriendsSerializer
from .models import Friends


# Create your views here.


class ListFriends(ListAPIView):
    serializer_class = FriendsSerializer
    authentication_classes = (TokenAuthentication,)

    # permission_classes = [IsAdminUser]

    def get_queryset(self):
        username = self.kwargs['username']
        friends = Friends.objects.list_friends_users(username)
        return friends


class ListFriends2(APIView):
    serializer_class = FriendsSerializer
    authentication_classes = (TokenAuthentication,)

    # permission_classes = [IsAdminUser]

    def get_object(self, username):
        return Friends.objects.list_friends_users(username)

    def get(self, request, username):
        friends = self.get_object(username)

        wb = Workbook()

        ws = wb.active
        ws["B2"] = "username"
        ws["C2"] = "name"
        ws["D2"] = "status"
        ws["E2"] = "city"
        i = 4
        for f in friends:
            ws.cell(row=i, column=2).value = f.name
            ws.cell(row=i, column=3).value = f.last_name
            ws.cell(row=i, column=4).value = f.status
            ws.cell(row=i, column=5).value = f.city
            i = i + 1

        file_name = username +"_friends.xlsx"

        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename={0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response

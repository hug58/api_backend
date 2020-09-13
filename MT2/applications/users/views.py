from django.http import HttpResponse
from rest_framework import status
from .serializers import UserSerializer, UserListSerializer
from rest_framework.views import APIView
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from applications.friends.models import Friends
from rest_framework.authtoken.models import Token
from .models import User
from openpyxl import Workbook


class Data(APIView):
    serializer_class = UserListSerializer

    def get(self, request):
        users = User.objects.all()
        wb = Workbook()
        ws = wb.active
        ws["B2"] = "username"
        ws["C2"] = "name"
        ws["D2"] = "last_name"
        ws["E2"] = "email"

        i = 4
        for user in users:
            ws.cell(row=i, column=2).value = user.username
            ws.cell(row=i, column=3).value = user.name
            ws.cell(row=i, column=4).value = user.last_name
            ws.cell(row=i, column=5).value = user.email
            i = i + 1

        file_name = "User.xlsx"

        response = HttpResponse(content_type="application/ms-excel")
        contain = "attachment; filename={0}".format(file_name)
        response["Content-Disposition"] = contain
        wb.save(response)
        return response


class RegisterUser(APIView):
    serializer_class = UserSerializer
    authentication_classes = (TokenAuthentication,)
    permission_classes = [IsAdminUser]

    def post(self, request):
        serializer = UserSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user, created = User.objects.get_or_create(
            username=serializer.validated_data['username'],
            name=serializer.validated_data['name'],
            last_name=serializer.validated_data['last_name'],
            email=serializer.validated_data['email'],
            is_active=True,
        )
        if created:

            token = Token.objects.create(user=user)
        else:

            token = Token.objects.get(user=user)

        user_get = {
            'id': user.pk,
            'username': user.username,
            'email': user.email,
            'name': user.name,
            'last_name': user.last_name,
        }

        friends = serializer.validated_data['friends']
        user_id = User.objects.get(id=user.pk)
        for friend in friends:
            Friends.objects.get_or_create(
                name=friend["name"],
                last_name=friend['last_name'],
                city=friend['city'],
                status=friend['status'],
                friend=user_id,
            )
        return Response({
            "status": 200,
            "message": "user and friends added",
            "token": token.key,
            "user": user_get,
            "friends": friends}, status=status.HTTP_201_CREATED)
